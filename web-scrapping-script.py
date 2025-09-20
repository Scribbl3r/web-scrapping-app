#!/usr/bin/env python
import requests
from bs4 import BeautifulSoup
import os
from pathlib import Path
from openpyxl import Workbook, load_workbook

headers = {
  'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/106.0.0.0 Safari/537.36',
  'Accept-Language': 'en-US,en;q=0.5'
}

# compatibility btw Python2 & Python3
_input = input
if hasattr(__builtins__, 'raw_input'):
    _input = raw_input

# GET INFO FROM 1 URL
def get_product_details(product_url: str) -> dict:
  # Create an empty product details dictionary
  product_details = {}

  # Get the product page content and create a soup
  page = requests.get(product_url, headers=headers)
  soup = BeautifulSoup(page.content, 'lxml')

  try :
    #title
    title = soup.find('span', attrs={'id':"productTitle"}).get_text().strip()
    #author
    author = soup.find('span', attrs={'class':"author"}).find_all("a")[0].get_text(strip=True, separator= " ")
    #note & reviews
    noteInfo = soup.find('span', attrs={'id':"acrPopover"}).find_all("i")[0]
    note = noteInfo.get_text().strip().split(" ")[0]
    reviews = soup.find('span', attrs={'id':"acrCustomerReviewText"}).get_text().strip()
    #publisher
    prodDetailsList = soup.find('ul', attrs={"class":'detail-bullet-list'}).find_all("li")[1]
    publisher = prodDetailsList.find_all('span')[2].text
    #rank PRBLM HERE
    allRankList = soup.find('ul', attrs={"class":'detail-bullet-list'}).find_all("li")
    rank =""
    for li in allRankList :
       if "Best Sellers Rank" in li.get_text():
            rankedList = li.find('ul')
            rank = rankedList.find_all('li')[0].text

    #add to product
    product_details["title"] = title
    product_details["author"] = author
    product_details["note"] = note
    product_details["reviews"] = reviews
    product_details["publisher"] = publisher
    product_details["rank"] = rank

    return product_details

  except Exception as e:
    print('Could not fetch product details')
    print(f'Failed with exception: {e}')

# PRINT results in txt file
def saveToTxt(object):
    with open("scrappEbook-results.txt", "a", encoding="utf-8") as file :
        file.write("==== Results for " + str(object.get("title", "N/A")) + " ====\n")
        file.write("titre : " + str(object.get("title", "N/A")) + "\n")
        file.write("author : " + str(object.get("author", "N/A")) + "\n")
        file.write("note : " + str(object.get("note", "N/A")) + "\n")
        file.write("reviews : " + str(object.get("reviews", "N/A")) + "\n")
        file.write("publisher : " + str(object.get("publisher", "N/A")) + "\n")
        file.write("rank : " + str(object.get("rank", "N/A")) + "\n")

# PRINT results in xlsx file
def saveToExcel(results, filename="scrappEbook-results.xlsx"):
    if os.path.exists(filename):
        # ouvrir fichier existant
        wb = load_workbook(filename)
        ws = wb.active
    else:
        # créer un nouveau fichier
        wb = Workbook()
        ws = wb.active
        ws.title = "Ebooks"
        # ajouter les en-têtes une seule fois
        headers = ["Title", "Author", "Note", "Reviews", "Publisher", "Rank"]
        ws.append(headers)

    # remplir les lignes
    for r in results:
        ws.append([
            r.get("title", "N/A"),
            r.get("author", "N/A"),
            r.get("note", "N/A"),
            r.get("reviews", "N/A"),
            r.get("publisher", "N/A"),
            r.get("rank", "N/A"),
        ])

    wb.save(filename)

# SCRAPP ALL
def scrapAllUrls():
    path = input("give me the path to your file with the Amazon URLs to scrap : ").strip('" \n\r\t')
    file_path = Path(path)
    with open(file_path, "r", encoding="utf-8") as f :
        content=f.read().strip()
        urls = [u.strip() for u in content.split(";") if u.strip()]

    info = []
    for url in urls :
       info.append(get_product_details(url))
       #saveToTxt(get_product_details(url))

    #print(info)
    saveToExcel(info)

    print("Results saved in txt file. Path :", os.path.abspath("scrappEbook-results.txt"))

scrapAllUrls()